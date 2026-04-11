#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║       CERTIFICATE AUTOMATION — FAST MODE (500+)            ║
║  Parallel processing + persistent SMTP = 10x faster        ║
╚══════════════════════════════════════════════════════════════╝

SETUP (run once):
    pip install pillow openpyxl reportlab tqdm python-dotenv

USAGE:
    python certificates_fast.py
"""

import os, sys, smtplib, json, time, threading
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

# ─────────────────────────────────────────────
#  🔐 LOAD CREDENTIALS FROM .env
# ─────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("\n❌ Missing package: pip install python-dotenv\n")
    sys.exit(1)

GMAIL_ADDRESS  = os.getenv("GMAIL_ADDRESS", "")
GMAIL_APP_PASS = os.getenv("GMAIL_APP_PASS", "")
EXCEL_FILE     = os.getenv("EXCEL_FILE",     "participants.xlsx")
CERTIFICATE_BG = os.getenv("CERTIFICATE_BG", "certificate_bg.png")
OUTPUT_FOLDER  = os.getenv("OUTPUT_FOLDER",  "certificates_output")

# ─────────────────────────────────────────────
#  ⚡ SPEED SETTINGS  (edit freely)
# ─────────────────────────────────────────────
WORKERS          = 10   # parallel certificate generators
SMTP_CONNECTIONS = 5    # parallel Gmail senders (keep ≤5)
BATCH_SIZE       = 50   # save progress every N certs

# ─────────────────────────────────────────────
#  📧 EMAIL CONTENT  (edit freely)
# ─────────────────────────────────────────────
EMAIL_SUBJECT = "🎓 Your Certificate of Participation – STRAT-A-THON 1.0"
EMAIL_BODY    = """Dear {name},

Congratulations! 🎉

We are delighted to present your Certificate of Participation for 
STRAT-A-THON 1.0, the 24-hour hackathon held on 4-5 March 2026 at 
Vishnu Institute of Technology, Bhimavaram.

Your active participation and exceptional performance made this event 
a great success. We are proud of your dedication and spirit!

Please find your personalized certificate attached to this email.

Best regards,
Techie Blazers Club
Department of CSBS
Vishnu Institute of Technology"""

# ─────────────────────────────────────────────
#  🎨 NAME POSITION ON CERTIFICATE  (edit freely)
# ─────────────────────────────────────────────
NAME_VERTICAL_POSITION   = 0.396  # pixel-exact for this certificate design
NAME_HORIZONTAL_POSITION = 0.50   # centered
FONT_SIZE                = 80     # fits cleanly in the gap
FONT_COLOR               = (255, 255, 255)  # white

# ─────────────────────────────────────────────
#  DO NOT EDIT BELOW THIS LINE
# ─────────────────────────────────────────────

PROGRESS_FILE = "progress.json"

# ── Dependency check ──────────────────────────────────────────
def check_dependencies():
    missing = []
    for pkg, imp in [("pillow", "PIL"), ("openpyxl", "openpyxl"),
                     ("reportlab", "reportlab"), ("tqdm", "tqdm")]:
        try:
            __import__(imp)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"\n❌ Missing packages: pip install {' '.join(missing)}\n")
        sys.exit(1)

def check_env():
    errors = []
    if not GMAIL_ADDRESS:
        errors.append("❌ GMAIL_ADDRESS is not set in .env")
    elif "@" not in GMAIL_ADDRESS:
        errors.append("❌ GMAIL_ADDRESS looks invalid in .env")
    if not GMAIL_APP_PASS:
        errors.append("❌ GMAIL_APP_PASS is not set in .env")
    elif GMAIL_APP_PASS == "xxxx xxxx xxxx xxxx":
        errors.append("❌ GMAIL_APP_PASS is still the placeholder — update .env")
    if not os.path.exists(EXCEL_FILE):
        errors.append(f"❌ Excel file not found: '{EXCEL_FILE}'")
    if not os.path.exists(CERTIFICATE_BG):
        errors.append(f"❌ Certificate background not found: '{CERTIFICATE_BG}'")
    if not os.path.exists(".env"):
        errors.append("⚠️  No .env file found — copy .env.example to .env and fill it in")
    if errors:
        print("\n" + "\n".join(errors))
        print("\n  👉 Copy .env.example → .env and fill in your values\n")
        sys.exit(1)
    print(f"  📧 Gmail  : {GMAIL_ADDRESS}")
    print(f"  📊 Excel  : {EXCEL_FILE}")
    print(f"  🎨 Design : {CERTIFICATE_BG}")

# ── Progress tracking (resume after crash) ────────────────────
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return set(json.load(f).get("sent", []))
    return set()

def save_progress(sent_emails):
    with open(PROGRESS_FILE, "w") as f:
        json.dump({"sent": list(sent_emails),
                   "last_updated": datetime.now().isoformat()}, f)

# ── Load participants ─────────────────────────────────────────
def load_participants():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    try:
        ni = headers.index("name")
        ei = headers.index("email")
    except ValueError:
        print(f"\n❌ Need columns 'Name' and 'Email'. Found: {headers}\n")
        sys.exit(1)
    people = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        n = str(row[ni]).strip() if row[ni] else ""
        e = str(row[ei]).strip() if row[ei] else ""
        if n and e and n != "None" and e != "None":
            people.append({"name": n, "email": e})
    return people

# ── Font loader (cached) ──────────────────────────────────────
_font_cache = {}
_font_lock  = threading.Lock()

def get_font(size):
    with _font_lock:
        if size in _font_cache:
            return _font_cache[size]
        from PIL import ImageFont
        paths = [
            "C:/Windows/Fonts/Garamond.ttf",
            "C:/Windows/Fonts/Georgia.ttf",
            "C:/Windows/Fonts/times.ttf",
            "C:/Windows/Fonts/Cambria.ttf",
            "C:/Windows/Fonts/arial.ttf",
            "/Library/Fonts/Georgia.ttf",
            "/System/Library/Fonts/Helvetica.ttc",
            "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
        ]
        font = None
        for p in paths:
            if os.path.exists(p):
                try:
                    font = ImageFont.truetype(p, size)
                    break
                except:
                    continue
        if not font:
            font = ImageFont.load_default()
        _font_cache[size] = font
        return font

# ── Background image (cached) ─────────────────────────────────
_bg_cache = None
_bg_lock  = threading.Lock()

def get_background():
    global _bg_cache
    with _bg_lock:
        if _bg_cache is None:
            from PIL import Image
            _bg_cache = Image.open(CERTIFICATE_BG).convert("RGBA")
        return _bg_cache.copy()

# ── Generate certificate PDF ──────────────────────────────────
def create_certificate_pdf(name):
    from PIL import Image, ImageDraw
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader

    bg = get_background()
    w, h = bg.size
    overlay = Image.new("RGBA", bg.size, (255, 255, 255, 0))
    draw    = ImageDraw.Draw(overlay)
    font    = get_font(FONT_SIZE)

    try:
        bbox   = draw.textbbox((0, 0), name, font=font)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
    except AttributeError:
        text_w, text_h = draw.textsize(name, font=font)

    x = int(w * NAME_HORIZONTAL_POSITION) - text_w // 2
    y = int(h * NAME_VERTICAL_POSITION)   - text_h // 2

    shadow = max(2, FONT_SIZE // 40)
    draw.text((x + shadow, y + shadow), name, font=font, fill=(0, 0, 0, 80))
    r, g, b = FONT_COLOR
    draw.text((x, y), name, font=font, fill=(r, g, b, 255))

    final   = Image.alpha_composite(bg, overlay).convert("RGB")
    pdf_buf = BytesIO()
    c       = canvas.Canvas(pdf_buf, pagesize=(841.89, 595.28))
    img_buf = BytesIO()
    final.save(img_buf, format="PNG")
    img_buf.seek(0)
    c.drawImage(ImageReader(img_buf), 0, 0, width=841.89, height=595.28)
    c.save()
    pdf_buf.seek(0)
    return pdf_buf

# ── Save PDF locally ──────────────────────────────────────────
def save_pdf(name, pdf_buf):
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    path = os.path.join(OUTPUT_FOLDER, f"Certificate_{name.replace(' ', '_')}.pdf")
    with open(path, "wb") as f:
        f.write(pdf_buf.read())
    return path

# ── SMTP connection pool ──────────────────────────────────────
class SMTPPool:
    def __init__(self, size):
        self._lock  = threading.Semaphore(size)
        self._local = threading.local()

    def _connect(self):
        conn = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        conn.login(GMAIL_ADDRESS, GMAIL_APP_PASS)
        return conn

    def send(self, msg):
        self._lock.acquire()
        try:
            if not hasattr(self._local, "conn") or self._local.conn is None:
                self._local.conn = self._connect()
            try:
                self._local.conn.send_message(msg)
            except smtplib.SMTPServerDisconnected:
                self._local.conn = self._connect()
                self._local.conn.send_message(msg)
        finally:
            self._lock.release()

smtp_pool = None

# ── Build email ───────────────────────────────────────────────
def build_email(name, email, pdf_buf):
    msg            = MIMEMultipart()
    msg["From"]    = GMAIL_ADDRESS
    msg["To"]      = email
    msg["Subject"] = EMAIL_SUBJECT.format(name=name)
    msg.attach(MIMEText(EMAIL_BODY.format(name=name), "plain"))
    att = MIMEApplication(pdf_buf.read(), _subtype="pdf")
    att.add_header("Content-Disposition", "attachment",
                   filename=f"Certificate_{name.replace(' ', '_')}.pdf")
    msg.attach(att)
    return msg

# ── Process one participant ───────────────────────────────────
def process_one(person, already_sent):
    name  = person["name"]
    email = person["email"]
    if email in already_sent:
        return {"status": "skipped", "name": name, "email": email}
    try:
        pdf_buf = create_certificate_pdf(name)
        save_pdf(name, pdf_buf)
        pdf_buf.seek(0)
        msg = build_email(name, email, pdf_buf)
        smtp_pool.send(msg)
        return {"status": "ok", "name": name, "email": email}
    except smtplib.SMTPAuthenticationError:
        return {"status": "auth_error", "name": name, "email": email}
    except Exception as e:
        return {"status": "error", "name": name, "email": email, "err": str(e)}

# ── Main ──────────────────────────────────────────────────────
def main():
    global smtp_pool

    print("\n╔══════════════════════════════════════════════╗")
    print("║   CERTIFICATE AUTOMATION — ⚡ FAST MODE     ║")
    print("╚══════════════════════════════════════════════╝\n")

    print("🔍 Checking dependencies...")
    check_dependencies()
    print("✅ Dependencies OK\n")

    print("🔐 Loading credentials from .env...")
    check_env()
    print("✅ Credentials OK\n")

    from tqdm import tqdm

    participants = load_participants()
    already_sent = load_progress()
    remaining    = [p for p in participants if p["email"] not in already_sent]

    print(f"\n📊 Total participants : {len(participants)}")
    print(f"✅ Already sent       : {len(already_sent)}")
    print(f"📬 To process now     : {len(remaining)}")
    print(f"⚡ Parallel workers   : {WORKERS}")
    print(f"📧 SMTP connections   : {SMTP_CONNECTIONS}")

    if not remaining:
        print("\n🎉 All certificates already sent!\n")
        sys.exit(0)

    est = len(remaining) * 0.3
    print(f"⏱️  Estimated time     : ~{int(est//60)}m {int(est%60)}s\n")

    confirm = input("▶  Start sending? (yes/no): ").strip().lower()
    if confirm not in ("yes", "y"):
        print("\n❌ Cancelled.\n")
        sys.exit(0)

    print(f"\n🚀 Starting at {datetime.now().strftime('%H:%M:%S')}...\n")

    smtp_pool   = SMTPPool(SMTP_CONNECTIONS)
    sent_set    = set(already_sent)
    success     = 0
    failed      = 0
    skipped     = 0
    fail_log    = []
    start_time  = time.time()
    batch_count = 0

    with ThreadPoolExecutor(max_workers=WORKERS) as executor:
        futures = {executor.submit(process_one, p, sent_set): p for p in remaining}
        with tqdm(total=len(remaining), desc="Sending", unit="cert",
                  bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]",
                  colour="green") as pbar:
            for future in as_completed(futures):
                result = future.result()
                if result["status"] == "ok":
                    success += 1
                    sent_set.add(result["email"])
                    batch_count += 1
                    if batch_count % BATCH_SIZE == 0:
                        save_progress(sent_set)
                elif result["status"] == "skipped":
                    skipped += 1
                elif result["status"] == "auth_error":
                    print(f"\n\n❌ Gmail authentication failed!")
                    print("   Check GMAIL_ADDRESS and GMAIL_APP_PASS in your .env file")
                    print("   Get App Password: myaccount.google.com → Security → App Passwords\n")
                    save_progress(sent_set)
                    sys.exit(1)
                else:
                    failed += 1
                    fail_log.append(f"{result['name']} ({result['email']}): {result.get('err', 'unknown')}")
                pbar.set_postfix({"✅": success, "❌": failed})
                pbar.update(1)

    save_progress(sent_set)
    elapsed = time.time() - start_time

    print(f"\n{'─'*50}")
    print(f"🎉 DONE in {int(elapsed//60)}m {int(elapsed%60)}s")
    print(f"   ✅ Sent       : {success}")
    print(f"   ⏭️  Skipped    : {skipped}  (already sent)")
    print(f"   ❌ Failed     : {failed}")
    print(f"   📁 PDFs saved : ./{OUTPUT_FOLDER}/")
    if fail_log:
        print(f"\n⚠️  Failed:")
        for f in fail_log:
            print(f"   • {f}")
        print("\n   Re-run to retry failed ones.\n")
    print(f"\n   Progress saved to: {PROGRESS_FILE}\n")

if __name__ == "__main__":
    main()
